# Before running, install Flask, openpyxl, and flask-cors:
# pip install Flask openpyxl Flask-Cors

from flask import Flask, request, send_file
from flask_cors import CORS # Import CORS
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
import base64
import io
from datetime import datetime
import re # Import regex module for filename sanitization

app = Flask(__name__)
CORS(app) # Enable CORS for all routes

# Helper function to format date-time for Excel (YYYY-MM-DD HH:MM:SS)
def format_datetime_for_excel(dt_string):
    if not dt_string:
        return ''
    try:
        # datetime-local input format is 'YYYY-MM-DDTHH:MM'
        dt_object = datetime.strptime(dt_string, '%Y-%m-%dT%H:%M')
        return dt_object.strftime('%Y-%m-%d %H:%M:%S')
    except ValueError:
        return dt_string # Return original if parsing fails

# Helper function to format date-time for safe filenames
def format_datetime_for_filename(dt_string):
    if not dt_string:
        return ''
    try:
        dt_object = datetime.strptime(dt_string, '%Y-%m-%dT%H:%M')
        # Format as YYYY-MM-DD_HHMM (safe for filenames)
        return dt_object.strftime('%Y-%m-%d_%H%M')
    except ValueError:
        return '' # Return empty string if parsing fails for filename

# Helper function to sanitize string for filename
def sanitize_filename(text):
    # Remove any characters that are not alphanumeric, underscores, or hyphens
    # Replace spaces with underscores
    text = text.replace(' ', '_')
    return re.sub(r'[^\w.-]', '', text)

@app.route('/generate_excel_with_images', methods=['POST'])
def generate_excel_with_images():
    """
    Receives incident report data and Base64 encoded images from the frontend,
    generates an Excel file with embedded images, and sends it back.
    """
    try:
        data = request.json
        if not data:
            return "No data received", 400

        # Extract all incident details and uploaded photos
        incident_details = data.get('incidentDetails', {})
        uploaded_photos = data.get('uploadedPhotos', [])

        # Create a new Excel workbook and a sheet for the report
        wb = Workbook()
        ws_report = wb.active
        ws_report.title = "Incident Report"

        # --- Populate Main Incident Report Sheet with data and formatting ---

        # Define styles
        bold_font = Font(bold=True)
        yellow_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")

        # Row 1: Title
        ws_report['A1'] = "B2B_Sub-Trunk_FTTx_MSAN uplink Access fiber Incident report template_LSP_MMP_June-2025_Sr_02"
        ws_report['A1'].font = bold_font
        ws_report['A1'].fill = yellow_fill
        ws_report.merge_cells('A1:F1') # Merge cells for the title

        # Row 2: Ticket Received Date Time, Customer Name
        ws_report['A2'] = "Ticket Received Date Time"
        ws_report['B2'] = format_datetime_for_excel(incident_details.get('ticketReceivedDateTime', ''))
        ws_report.merge_cells('B2:C2')
        ws_report['D2'] = "Customer Name"
        ws_report['E2'] = incident_details.get('customerName', '')
        ws_report.merge_cells('E2:F2')

        # Row 3: Circuit ID, Customer Address
        ws_report['A3'] = "Circuit ID"
        ws_report['B3'] = incident_details.get('circuitID', '')
        ws_report.merge_cells('B3:C3')
        ws_report['D3'] = "Customer Address"
        ws_report['E3'] = incident_details.get('customerAddress', '')
        ws_report.merge_cells('E3:F3')

        # Row 4: Type of Reaction, Work Order Email Title
        ws_report['A4'] = "Type of Reaction"
        ws_report['B4'] = incident_details.get('typeOfReaction', '')
        ws_report.merge_cells('B4:C4')
        ws_report['D4'] = "Work Order Email Title"
        ws_report['E4'] = incident_details.get('workOrderEmailTitle', '')
        ws_report.merge_cells('E4:F4')

        # Row 5: Empty row for spacing (no content, just for structure)
        # ws_report.merge_cells('A5:F5') # Not strictly needed if empty, but for consistency

        # Row 6: Table header (Description, Start Time, End Time, Action By LSP)
        ws_report['A6'] = "Description"
        ws_report['B6'] = "Start Time"
        ws_report['C6'] = "End Time"
        ws_report['D6'] = "Action By LSP"
        ws_report.merge_cells('D6:F6') # Merge cells for "Action By LSP" header

        # Row 7: WO Start
        ws_report['A7'] = "WO start"
        ws_report['B7'] = format_datetime_for_excel(incident_details.get('woStartTime', ''))
        ws_report['C7'] = format_datetime_for_excel(incident_details.get('woEndTime', ''))
        ws_report['D7'] = incident_details.get('woActionByLSP', '')
        ws_report.merge_cells('D7:F7')

        # Row 8: Arrived at Customer Premise/Exchange/RSU/BTS
        ws_report['A8'] = "Arrived at Customer Premise/Exchange/RSU/BTS from ……….."
        ws_report['B8'] = format_datetime_for_excel(incident_details.get('arrivedStartTime', ''))
        ws_report['C8'] = format_datetime_for_excel(incident_details.get('arrivedEndTime', ''))
        ws_report['D8'] = incident_details.get('arrivedActionByLSP', '')
        ws_report.merge_cells('D8:F8')

        # Row 9: Power meter/OTDR testing
        ws_report['A9'] = "Power meter/OTDR testing from customer/Exchange/RSU/BTS"
        ws_report['B9'] = format_datetime_for_excel(incident_details.get('powerMeterStartTime', 'N/A'))
        ws_report['C9'] = format_datetime_for_excel(incident_details.get('powerMeterEndTime', 'N/A'))
        ws_report['D9'] = incident_details.get('powerMeterActionByLSP', 'N/A')
        ws_report.merge_cells('D9:F9')

        # Row 10: Cable damage/cut distance
        ws_report['A10'] = "Cable damage/cut distance(Meter or km) from customer/Exchange/ Customer Site according to OTDR Test"
        ws_report['B10'] = format_datetime_for_excel(incident_details.get('cableDamageStartTime', 'N/A'))
        ws_report['C10'] = format_datetime_for_excel(incident_details.get('cableDamageEndTime', 'N/A'))
        ws_report['D10'] = incident_details.get('cableDamageActionByLSP', 'N/A')
        ws_report.merge_cells('D10:F10')

        # Row 11: Root Cause (Direct Affect) - 1
        ws_report['A11'] = "Root Cause (Direct affect)"
        ws_report['B11'] = format_datetime_for_excel(incident_details.get('rootCause1StartTime', ''))
        ws_report['C11'] = format_datetime_for_excel(incident_details.get('rootCause1EndTime', ''))
        ws_report['D11'] = incident_details.get('rootCause1ActionByLSP', '')
        ws_report.merge_cells('D11:F11')

        # Row 12: Rectification
        ws_report['A12'] = "Rectification"
        ws_report['B12'] = format_datetime_for_excel(incident_details.get('rectificationStartTime', 'N/A'))
        ws_report['C12'] = format_datetime_for_excel(incident_details.get('rectificationEndTime', ''))
        ws_report['D12'] = incident_details.get('rectificationActionByLSP', '')
        ws_report.merge_cells('D12:F12')

        # Row 13: Ping test and Speed test (If needed)
        ws_report['A13'] = "Ping test and Speed test(If needed)"
        ws_report['B13'] = format_datetime_for_excel(incident_details.get('pingTestStartTime', 'N/A'))
        ws_report['C13'] = format_datetime_for_excel(incident_details.get('pingTestEndTime', 'N/A'))
        ws_report['D13'] = incident_details.get('pingTestActionByLSP', 'N/A')
        ws_report.merge_cells('D13:F13')

        # Row 14: Service recovery confirmed by TSC/FSC
        ws_report['A14'] = "Service recovery confirmed by TSC/FSC"
        ws_report['B14'] = format_datetime_for_excel(incident_details.get('tscFscStartTime', 'N/A'))
        ws_report['C14'] = format_datetime_for_excel(incident_details.get('tscFscEndTime', ''))
        ws_report['D14'] = incident_details.get('tscFscActionByLSP', '')
        ws_report.merge_cells('D14:F14')

        # Row 15: Service recovery confirmed by Customer
        ws_report['A15'] = "Service recovery confirmed by Customer"
        ws_report['B15'] = format_datetime_for_excel(incident_details.get('customerConfirmStartTime', 'N/A'))
        ws_report['C15'] = format_datetime_for_excel(incident_details.get('customerConfirmEndTime', ''))
        ws_report['D15'] = incident_details.get('customerConfirmActionByLSP', '')
        ws_report.merge_cells('D15:F15')

        # Row 16: Outage duration
        ws_report['A16'] = "Outage duration"
        ws_report['B16'] = format_datetime_for_excel(incident_details.get('outageDurationStartTime', ''))
        ws_report['C16'] = format_datetime_for_excel(incident_details.get('outageDurationEndTime', ''))
        ws_report['D16'] = incident_details.get('outageDurationActionByLSP', '')
        ws_report.merge_cells('D16:F16')

        # Row 17: Empty row for spacing
        # ws_report.merge_cells('A17:F17') # Not strictly needed if empty

        # Row 18: GPS Location for Pole, GPS Location for Joint Closure
        ws_report['A18'] = "GPS Location for Pole( if replacement or new)"
        ws_report['B18'] = incident_details.get('gpsLocationPole', '')
        ws_report.merge_cells('B18:C18')
        ws_report['D18'] = "GPS Location for Joint Closure( if replacement or new)"
        ws_report['E18'] = incident_details.get('gpsLocationJointClosure', '')
        ws_report.merge_cells('E18:F18')

        # Row 19: Media Converter/ONT/IAD, OTDR test result
        ws_report['A19'] = "Media Converter/ONT/IAD equipment old&new serial number if replaceed"
        ws_report['B19'] = incident_details.get('mediaConverterSerial', '')
        ws_report.merge_cells('B19:C19')
        ws_report['D19'] = "OTDR test result(to provide in separate sheet and (PDF or SOR file))"
        ws_report['E19'] = incident_details.get('otdrTestResultNotes', '')
        ws_report.merge_cells('E19:F19')

        # Row 20: Sub-Root cause (External Affect), Root Cause (Direct affect) - 2
        ws_report['A20'] = "Sub-Root cause (External Affect)"
        ws_report['B20'] = incident_details.get('subRootCauseExternalAffect', '')
        ws_report.merge_cells('B20:C20')
        ws_report['D20'] = "Root Cause (Direct affect)"
        ws_report['E20'] = incident_details.get('rootCauseDirectAffect2', '')
        ws_report.merge_cells('E20:F20')

        # Row 21: Additional Notes
        ws_report['A21'] = "" # This was empty in your JS, so keeping it
        ws_report.merge_cells('A21:C21') # Merge cells for "Additional Notes" label
        ws_report['D21'] = incident_details.get('additionalNotes', '')
        ws_report.merge_cells('D21:F21') # Merge cells for "Additional Notes" value

        # Set column widths
        ws_report.column_dimensions['A'].width = 45
        ws_report.column_dimensions['B'].width = 20
        ws_report.column_dimensions['C'].width = 20
        ws_report.column_dimensions['D'].width = 30
        ws_report.column_dimensions['E'].width = 30
        ws_report.column_dimensions['F'].width = 20


        # --- Create a new sheet for Photos with Embedded Images ---
        ws_photos = wb.create_sheet("Embedded Photos")
        ws_photos['A1'] = "Photo Label"
        ws_photos['B1'] = "Image"
        ws_photos.column_dimensions['A'].width = 40 # Set width for label column
        ws_photos.column_dimensions['B'].width = 50 # Set width for image column (adjust as needed)

        row_num = 2 # Start from row 2 for photo data

        for photo in uploaded_photos:
            label = photo.get('label', 'No Label')
            base64_data = photo.get('data', '')

            # Write the label to column A
            ws_photos[f'A{row_num}'] = label

            if base64_data:
                try:
                    # Remove the data URL prefix (e.g., "data:image/png;base64,")
                    if ',' in base64_data:
                        base64_decoded_string = base64_data.split(',')[1]
                    else:
                        base64_decoded_string = base64_data

                    # Decode Base64 string to bytes
                    img_bytes = base64.b64decode(base64_decoded_string)
                    img_stream = io.BytesIO(img_bytes)

                    # Create an openpyxl Image object
                    # openpyxl usually infers from the bytes, but explicit is safer for some cases.
                    img = Image(img_stream)

                    # Set image dimensions (optional, but good for consistent display)
                    # You may want to resize images on the client-side before sending
                    # or implement more sophisticated resizing here.
                    img.width = 300 # Example width
                    img.height = 200 # Example height

                    # Add the image to the worksheet, anchored to cell B (where the image will appear)
                    ws_photos.add_image(img, f'B{row_num}')

                    # Adjust row height to fit the image
                    # Approx conversion from pixels to Excel row height units (1px ~ 0.75 Excel units)
                    ws_photos.row_dimensions[row_num].height = img.height * 0.75

                except Exception as img_error:
                    print(f"Error processing image {photo.get('name')}: {img_error}")
                    ws_photos[f'B{row_num}'] = f"Error embedding image: {img_error}"
            else:
                ws_photos[f'B{row_num}'] = "No image data"

            row_num += 1

        # Determine the dynamic filename
        customer_name = sanitize_filename(incident_details.get('customerName', ''))
        ticket_time = format_datetime_for_filename(incident_details.get('ticketReceivedDateTime', ''))
        
        if customer_name and ticket_time:
            download_filename = f"{customer_name}_{ticket_time}_IncidentReport.xlsx"
        elif customer_name:
            download_filename = f"{customer_name}_IncidentReport.xlsx"
        elif ticket_time:
            download_filename = f"IncidentReport_{ticket_time}.xlsx"
        else:
            download_filename = "incident_report_with_photos.xlsx"


        # Save the workbook to a BytesIO object
        excel_file_stream = io.BytesIO()
        wb.save(excel_file_stream)
        excel_file_stream.seek(0) # Rewind to the beginning of the stream

        # Send the file as a response
        return send_file(
            excel_file_stream,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=download_filename # Use the dynamic filename here
        )

    except Exception as e:
        print(f"An error occurred: {e}")
        return f"An internal server error occurred: {e}", 500

if __name__ == '__main__':
    app.run(debug=True)