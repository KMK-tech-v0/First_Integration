import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_standard_template():
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========== Template Sheet ==========
    template_sheet = wb.create_sheet("Template")
    
    # Hide gridlines (correct way for openpyxl 3.0+)
    template_sheet.sheet_view.showGridLines = False
    
    # Set column widths
    template_sheet.column_dimensions['A'].width = 40
    template_sheet.column_dimensions['B'].width = 25
    template_sheet.column_dimensions['C'].width = 20
    template_sheet.column_dimensions['D'].width = 50
    
    # Header style
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # Title row
    template_sheet.merge_cells('A1:D1')
    title_cell = template_sheet['A1']
    title_cell.value = "B2B_Sub-Trunk_FTTx_MSAN uplink Access fiber Incident report template_LSP_MMP_Feb-2025_Sr.05"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Incident Information section
    incident_info = [
        ["Ticket Received Date Time", "", "", ""],
        ["Customer Name", "FTTH pole & Cable damages in Swebo", "", ""],
        ["Circuit ID", "Referred to mail", "", ""],
        ["Customer Address", "N/A", "", ""],
        ["Type of Reaction", "", "", ""],
        ["Work Order Email Title", "FTTH pole & Cable damages in Swebo", "", ""],
        ["", "", "", ""]
    ]
    
    for row_idx, row in enumerate(incident_info, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = template_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 2 or row_idx == 8:  # Header and empty row
                cell.fill = header_fill
                cell.font = header_font
    
    # Timeline section header
    timeline_header = ["Description", "Start Time", "End Time", "Action By LSP"]
    for col_idx, header in enumerate(timeline_header, start=1):
        cell = template_sheet.cell(row=9, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Timeline data
    timeline_data = [
        ["WO start", "2024-06-14 09:32:00", "", "Mail/SMS Received"],
        ["Arrived at Customer Premise/Exchange/RSU/BTS from ………..", "2024-06-15 10:30:00", "2025-02-01 10:12:00", "Arrive fault location"],
        ["Power meter/OTDR testing from customer/Exchange/RSU/BTS", "N/A", "N/A", "N/A"],
        ["Cable damage/cut distance(Meter or km) from customer/Exchange/Customer Site according to OTDR Test", "N/A", "N/A", "N/A"],
        ["Root Cause (Direct affect)", "2024-06-15 10:30:00", "2025-02-01 10:12:00", "Cable tension"],
        ["Rectification", "N/A", "2025-02-01 10:12:00", "already done"],
        ["Ping test and Speed test(If needed)", "N/A", "N/A", "N/A"],
        ["Service recovery confirmed by TSC/FSC", "N/A", "2025-02-01 10:12:00", "already done"],
        ["Service recovery confirmed by Customer", "N/A", "2025-02-01 10:12:00", "already done"],
        ["Outage duration", "2024-06-14 09:32:00", "2025-02-01 10:12:00", "=C19-B19"],
        ["", "", "", ""],
        ["GPS Location for Pole(if replacement or new)", "", "", "N/A"],
        ["GPS Location for Joint Closure(if replacement or new)", "", "", "N/A"],
        ["Media Converter/ONT/IAD equipment old&new serial number if replaced", "", "", "N/A"],
        ["OTDR test result(to provide in separate sheet and (PDF or SOR file))", "", "", "N/A"],
        ["", "", "", ""],
        ["Sub-Root cause (External Affect)", "", "", ""],
        ["Root Cause (Direct affect)", "", "", "Pole broken down due to cable tension"],
        ["Way of recovery", "", "", "MMP team firstly check the fault point and pole broken due to cable tension. Team replaced new 8m 2 poles and link was restored."],
        ["Remark: Time delay due to difficult to buy poles.", "", "", ""]
    ]
    
    for row_idx, row in enumerate(timeline_data, start=10):
        for col_idx, value in enumerate(row, start=1):
            cell = template_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx in [10, 20, 27]:  # Highlight certain rows
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # ========== Photos Sheet ==========
    photos_sheet = wb.create_sheet("Photos")
    photos_sheet.sheet_view.showGridLines = False
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        photos_sheet.column_dimensions[col].width = 15
    
    # Set row heights for photo areas
    photo_row_height = 75  # Height for rows with photos
    
    # Photo sections - we'll create 10 photo positions as requested
    photo_sections = [
        ("Photos of OTDR or OPM Test Result", 1),
        ("Photos of Fault Location", 2),
        ("Photos of Rectification", 3),
        ("Photos of Speed Test/Ping Test", 4),
        ("Photos of Equipment", 5),
        ("Photos of Cable Damage", 6),
        ("Photos of Pole Replacement", 7),
        ("Photos of Site Overview", 8),
        ("Photos of Testing Equipment", 9),
        ("Photos of Final Check", 10)
    ]
    
    current_row = 1
    for section, num in photo_sections:
        # Section header
        photos_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=16)
        header_cell = photos_sheet.cell(row=current_row, column=1, value=f"Description: {section}")
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        current_row += 1
        
        # Sub-header
        photos_sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
        photos_sheet.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=10)
        photos_sheet.merge_cells(start_row=current_row, start_column=12, end_row=current_row, end_column=15)
        
        before_cell = photos_sheet.cell(row=current_row, column=2, value="Before")
        during_cell = photos_sheet.cell(row=current_row, column=7, value="During")
        after_cell = photos_sheet.cell(row=current_row, column=12, value="After")
        
        for cell in [before_cell, during_cell, after_cell]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Photo placeholders (3 rows for each photo type)
        for photo_type, col in [("Before", 2), ("During", 7), ("After", 12)]:
            # Set row height for photo area
            for i in range(3):
                photos_sheet.row_dimensions[current_row + i].height = photo_row_height
            
            # Merge cells for photo placeholder
            photos_sheet.merge_cells(start_row=current_row, start_column=col, 
                                    end_row=current_row+2, end_column=col+3)
            
            # Add placeholder text
            placeholder = photos_sheet.cell(row=current_row, column=col, 
                                          value=f"[{photo_type} Photo Placeholder\nSection {num}]")
            placeholder.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            placeholder.font = Font(color="808080", italic=True)
            
            # Add borders
            for r in range(current_row, current_row+3):
                for c in range(col, col+4):
                    photos_sheet.cell(row=r, column=c).border = thin_border
        
        current_row += 3
        
        # Add remark
        photos_sheet.cell(row=current_row, column=1, value="No")
        photos_sheet.cell(row=current_row, column=2, value=num)
        photos_sheet.cell(row=current_row, column=3, value="Date")
        photos_sheet.cell(row=current_row, column=4, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        photos_sheet.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
        photos_sheet.cell(row=current_row, column=6, value="Remark")
        photos_sheet.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=16)
        
        current_row += 2
    
    # Add final remark
    photos_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=16)
    photos_sheet.cell(row=current_row, column=1, value="Remark: Evident Photos have to be taken before repairing work")
    photos_sheet.cell(row=current_row, column=1).font = Font(italic=True)
    photos_sheet.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
    
    # Save the workbook
    filename = f"Standardized_Incident_Report_Template_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    print(f"Template created successfully: {filename}")

if __name__ == "__main__":
    create_standard_template()