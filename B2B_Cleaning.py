from openpyxl import load_workbook
import re
import datetime
import pandas as pd

file_path = r'D:\My Base\Share_Analyst\B2B\ME\To Do\B2B_summary.xlsx'
sheet_name = 'Records'
df = pd.read_excel(file_path, sheet_name=sheet_name)
df = pd.DataFrame(df)
df_col = df.columns.tolist()
print(df_col)
# Columns you want to clean (case-insensitive, spaces normalized)
target_columns = [
    "CASE TITLE",
    "CIRCUIT ID",
    "SERVICE TERMINATION POINT",
    "ADDRESS",
    "Status",
    "STATE/REGION",
    "TOWNSHIP",
    "TICKET STATUS",
    "SUB-ROOT CAUSE (EXTERNAL AFFECT)",
    "ROOT CAUSE (DIRECT AFFECT)",
    "ACTION TAKEN",
    "MATERIAL REPLACEMENT (IF HAS)",
    "TYPE OF AFFECTED SERVICE",
    "WORK ORDER FROM(FSC/TSC/FM1 ETC...)",
    "PIC",
    "DT_RANGE",
    "REPORTED"
]

def normalize_header(text):
    return text.strip().upper().replace('\xa0', ' ')  # normalize spaces & case

wb = load_workbook(file_path)
ws = wb[sheet_name]

# Read header row and map normalized header → column index (0-based)
header_row = [normalize_header(cell.value or "") for cell in ws[1]]

# Find column indexes of target columns present in the sheet
target_col_indexes = []
for col_name in target_columns:
    try:
        idx = header_row.index(normalize_header(col_name))
        target_col_indexes.append(idx)
    except ValueError:
        print(f"⚠️ Column '{col_name}' not found in sheet headers")

# Detect string columns among target columns (sample first 10 rows)
string_cols = set()
sample_rows = min(10, ws.max_row - 1)

for col_idx in target_col_indexes:
    for row in range(2, 2 + sample_rows):
        val = ws.cell(row=row, column=col_idx + 1).value
        if isinstance(val, str):
            string_cols.add(col_idx)
            break

# Clean function
def clean_text(value):
    if not isinstance(value, str):
        return value
    cleaned = re.sub(r'[\x00-\x1F]+', '', value).strip()
    return cleaned.upper()

# Apply cleaning to string cells in target columns only
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for col_idx in string_cols:
        cell = row[col_idx]
        cell.value = clean_text(cell.value)

# Find the latest complaint date
complaint_col_idx = header_row.index(normalize_header("COMPLAINT ISSUE TIME"))
complaint_dates = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    val = row[complaint_col_idx].value
    dt = None
    if isinstance(val, (datetime.datetime, datetime.date)):
        dt = val.date() if isinstance(val, datetime.datetime) else val
    else:
        try:
            dt = datetime.datetime.strptime(str(val), "%Y-%m-%d").date()
        except Exception:
            pass
    if dt:
        complaint_dates.append(dt)

if complaint_dates:
    last_date = max(complaint_dates)
else:
    last_date = datetime.date(2025, 6, 27)  # fallback

# Week definition: each week is Friday to Thursday (inclusive)
# Previous week: 13th June 2025 (Friday) to 19th June 2025 (Thursday)
# So, week starts on Friday, ends on Thursday

# Find the first Friday on/after your start date
start_date = datetime.date(2022, 4, 2)
while start_date.weekday() != 4:  # 4 = Friday
    start_date += datetime.timedelta(days=1)

week_ranges = []
current_start = start_date
while current_start <= last_date:
    current_end = current_start + datetime.timedelta(days=6)  # Thursday
    week_ranges.append((current_start, current_end))
    current_start = current_end + datetime.timedelta(days=1)

# Find or create REPORT_IN column
try:
    report_in_col_idx = header_row.index(normalize_header("REPORT_IN"))
except ValueError:
    ws.cell(row=1, column=ws.max_column + 1, value="REPORT_IN")
    report_in_col_idx = ws.max_column - 1  # 0-based

def parse_excel_date(val):
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.date() if isinstance(val, datetime.datetime) else val
    try:
        return datetime.datetime.strptime(str(val), "%Y-%m-%d").date()
    except Exception:
        return None

# Assign week range to REPORT_IN
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    cell_val = row[complaint_col_idx].value
    date_val = parse_excel_date(cell_val)
    week_label = ""
    if date_val:
        for i, (w_start, w_end) in enumerate(week_ranges):
            if w_start <= date_val <= w_end:
                week_label = f"Week {i+1}: {w_start} to {w_end}"
                break
    row[report_in_col_idx].value = week_label

# Save changes
wb.save(file_path)
print("✅ Excel updated: CLEAN + TRIM + selected UPPER done.")